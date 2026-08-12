#!/usr/bin/env python3
"""从教务导出的全校课程表生成地图地点回归测试数据。

脚本只读取“学校校区”和“已排时间地点”，按 (校区, 原始地点) 去重；不会把
课程、教师或学生相关字段写入仓库。已有 fixture 中人工确认过的 expected 会被
保留。发现新地点时默认拒绝覆盖，避免把尚未审核的地点误当成“不应显示 pin”。
"""

from __future__ import annotations

import argparse
import json
import re
import warnings
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as error:
    raise SystemExit("缺少 openpyxl，请运行：python3 -m pip install openpyxl") from error

SUPPORTED_CAMPUSES = {"中关村校区", "良乡校区"}
TIME_PATTERN = re.compile(
    r"星期[一二三四五六日天]\s*"
    r"(?:\[[^\]]+\]|第\s*\d+\s*节(?:\s*-\s*第\s*\d+\s*节)?)"
    r"\s*(.*?)"
    r"(?=(?:[;；]|[,，]\s*(?=[0-9单双周第()（）\-—至~、，,\s]*星期))|$)"
)


def extract_unique_locations(path: Path) -> list[tuple[str, str]]:
    warnings.filterwarnings("ignore", message="Workbook contains no default style")
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()  # 教务导出偶尔错误地把 dimension 标成 A1。
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    columns = {name: index for index, name in enumerate(headers)}
    required = {"学校校区", "已排时间地点"}
    if not required.issubset(columns):
        raise ValueError(f"工作簿缺少列：{', '.join(sorted(required - columns.keys()))}")

    locations: set[tuple[str, str]] = set()
    for row in rows:
        campus = str(row[columns["学校校区"]] or "").strip()
        schedule = row[columns["已排时间地点"]]
        if campus not in SUPPORTED_CAMPUSES or not schedule:
            continue
        for match in TIME_PATTERN.finditer(str(schedule)):
            classroom = match.group(1).strip(" \t,，;；")
            if classroom:
                locations.add((campus, classroom))
    return sorted(locations)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument(
        "--fixture",
        type=Path,
        default=Path("BIT101-iOSTests/Fixtures/campus-map-locations.json"),
    )
    parser.add_argument(
        "--allow-new",
        action="store_true",
        help="将尚未审核的新地点以 expected=null 写入（通常不应在 CI 使用）",
    )
    args = parser.parse_args()

    old_rows = json.loads(args.fixture.read_text(encoding="utf-8")) if args.fixture.exists() else []
    decisions = {(row["campus"], row["classroom"]): row.get("expected") for row in old_rows}
    locations = extract_unique_locations(args.xlsx)
    new_locations = [item for item in locations if item not in decisions]
    if new_locations and not args.allow_new:
        preview = "\n".join(f"- {campus} / {classroom}" for campus, classroom in new_locations[:20])
        raise SystemExit(
            f"发现 {len(new_locations)} 个尚未审核的新地点，fixture 未改动：\n{preview}\n"
            "审核后更新 expected，或显式使用 --allow-new 生成待处理项。"
        )

    rows = [
        {"campus": campus, "classroom": classroom, "expected": decisions.get((campus, classroom))}
        for campus, classroom in locations
    ]
    args.fixture.parent.mkdir(parents=True, exist_ok=True)
    args.fixture.write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"写入 {len(rows)} 条去重地点；新增待审核 {len(new_locations)} 条：{args.fixture}")


if __name__ == "__main__":
    main()
